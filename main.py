import openpyxl
import os
from openpyxl import load_workbook
from matplotlib import pyplot as plt
from openpyxl.drawing.image import Image
import numpy as np
from scipy.stats import skew
from statsmodels.graphics.tsaplots import plot_acf
import pandas as pd


# def intervals():
#    intervals = [1, 21, 63, 126, 252]
#    print(
#        "Daily, monthly, three-monthly, and annual volatility values have been added to the table. If you wish to add more values, enter +")
#    ind = input()
#    while ind == '+':
#        print("Enter the length of the interval(number of trading days)")
#        interval = int(input())
#        intervals.append(interval)
#        print("Value have been added to the table. If you wish to add more values, enter +")
#        ind = input()
#    return intervals

def count_volatility(dframe_fund,col_name):
   volatility_day = dframe_fund[col_name].std()
   volatility = []
   for index in range(len(intervals)):
       volatility_interval = volatility_day * np.sqrt(intervals[index])
       volatility.append(volatility_interval)
   return volatility

def rolling_std_df(df,col_name, window):
   df['rolling_vol ' + col_name] = df[col_name].rolling(window).std()*(252**0.5)
   return df

def fund_years(list, window):
   df_file = {}
   for index in range(len(list)):
       df_file[index] = pd.read_excel(list[index])
   df_append = pd.concat(df_file)

   df_2022 = pd.read_excel(list[0])
   df_agg = df_2022[['dt','ctrb_day']].groupby('dt').sum().reset_index()
   df_agg['day_week'] = pd.to_datetime(df_agg['dt']).dt.dayofweek
   df_agg.drop(df_agg[df_agg['day_week'] == 5].index, inplace=True)
   df_agg.drop(df_agg[df_agg['day_week'] == 6].index, inplace=True)

   df_append['day_week'] = pd.to_datetime(df_append['dt']).dt.dayofweek
   df_append.drop(df_append[df_append['day_week'] == 5].index, inplace=True)
   df_append.drop(df_append[df_append['day_week'] == 6].index, inplace=True)
   df_append = df_append.sort_values(by='dt')

   dframe_fund = df_append.groupby(['dt']).agg({'ctrb_day':'sum'}).reset_index()

   volatility = count_volatility(dframe_fund, 'ctrb_day')
   df_rolling = rolling_std_df(dframe_fund, 'ctrb_day', window)
   df_rolling = df_rolling.drop(columns='ctrb_day')

   volatility_name = ["daily", "month(21 trading days)", "3 month(63 trading days)", "6 month (63 trading days)", "year(252 trading days)"]
   for index in range(len(intervals) - 5):
       index = index + 5
       volatility_name.append(str(intervals[index]) + " trading days")

   df_volatility = {'Interval': volatility_name, 'Volatility': volatility}
   dframe_volatility = pd.DataFrame(df_volatility)

   # create an empty sheet, write the dataframe into it
   FilePath = "volatility_fund_years.xlsx"
   writer = pd.ExcelWriter(FilePath, engine='xlsxwriter')
   dframe_fund.to_excel(writer, sheet_name='fund_portfolio level', startrow=0, startcol=0)
   dframe_volatility.to_excel(writer, sheet_name='fund_portfolio level', startrow=0, startcol=5)
   writer.save()

   str_vol = str(round(volatility[4], 5) * 100)
   plt.hist(dframe_fund['ctrb_day'], bins=100, alpha=0.6)
   plt.title("Volatility " + str_vol + "%")
   plt.xlabel("Return")
   plt.xlabel("Frequency of return")
   plt.savefig('Frequency of return.png')
   plt.show()

   plt.plot(df_rolling['dt'], df_rolling['rolling_vol'])
   plt.title("Rolling volatility " + str(window) +" days")
   plt.savefig('rolling volatility_fund.png')
   plt.show()

   dframe_fund['Sum120'] = dframe_fund['ctrb_day'].rolling(120).sum()
   plt.plot(dframe_fund['dt'], dframe_fund['Sum120'])
   skew1 = str(round(skew(dframe_fund['Sum120'].dropna()), 5))
   plt.title("Skewness for data " + skew1)
   plt.savefig('Skewness for data.png')
   plt.show()

   ctrb_fund_1 = dframe_fund['ctrb_day'].tolist()
   ctrb_fund_1_array = np.asarray(ctrb_fund_1)

   plt.title("Autocorrelation Plot")
   plt.xlabel("Lags")
   plt.acorr(ctrb_fund_1_array, maxlags=None)
   plt.grid(True)
   plt.savefig('Autocorrelation_plot_2.png')
   plt.show()

   plot_acf(ctrb_fund_1_array).savefig('Autocorrelation_plot_1.png')

   wb = openpyxl.load_workbook(FilePath)
   wb.create_sheet('fund_portfolio_graph')
   active = wb['fund_portfolio_graph']
   active.add_image(Image('Frequency of return.png'), 'A1')
   active.add_image(Image('rolling volatility_fund.png'), 'K1')
   active.add_image(Image('Skewness for data.png'), 'K27')
   active.add_image(Image('Autocorrelation_plot_1.png'), 'U1')
   active.add_image(Image('Autocorrelation_plot_2.png'), 'U27')
   wb.save(FilePath)

   os.remove('Frequency of return.png')
   os.remove('rolling volatility_fund.png')
   os.remove('Skewness for data.png')
   os.remove('Autocorrelation_plot_1.png')
   os.remove('Autocorrelation_plot_2.png')

def fund_year(File):
   df_file = pd.read_excel(File)
   df_file['day_week'] = pd.to_datetime(df_file['dt']).dt.dayofweek
   df_file.drop(df_file[df_file['day_week'] == 5].index, inplace=True)
   df_file.drop(df_file[df_file['day_week'] == 6].index, inplace=True)
   df_file = df_file.sort_values(by='dt')
   File = '2017'

   volatility_name = ["daily", "month (21 trading days)", "3 month (63 trading days)", "6 month (63 trading days)",
                      "year (252 trading days)"]
   for index in range(len(intervals) - 5):
       index = index + 5
       volatility_name.append(str(intervals[index]) + " trading days")
   df_info_vol = pd.DataFrame(volatility_name)

   def volatility_portfolio():
       dframe_fund = df_file.groupby(['dt']).agg({'ctrb_day': 'sum'}).reset_index()

       volatility = count_volatility(dframe_fund, 'ctrb_day')

       volatility_name = ["daily", "month(21 trading days)", "3 month(63 trading days)", "6 month (63 trading days)",
                          "year(252 trading days)"]
       for index in range(len(intervals) - 5):
           index = index + 5
           volatility_name.append(str(intervals[index]) + " trading days")

       df_volatility = {'Interval': volatility_name, 'Volatility': volatility}
       dframe_volatility = pd.DataFrame(df_volatility)

       df_rolling = rolling_std_df(dframe_fund, 'ctrb_day', window)
       df_rolling = df_rolling.drop(columns='ctrb_day')

       # create an empty sheet, write the dataframe into it
       FilePath = "volatility_fund " + File + ".xlsx"
       writer = pd.ExcelWriter(FilePath, engine='xlsxwriter')
       df_rolling.to_excel(writer, sheet_name='fund_portfolio level', startrow=0, startcol=0)
       dframe_volatility.to_excel(writer, sheet_name='fund_portfolio level', startrow=0, startcol=5)
       writer.save()

       str_vol = str(round(volatility[4], 5) * 100)
       plt.hist(dframe_fund['ctrb_day'], bins=100, alpha=0.6)
       plt.title("Volatility " + str_vol + "%")
       plt.xlabel("Return")
       plt.xlabel("Frequency of return")
       plt.savefig('Frequency of return.png')
       plt.show()

       plt.plot(df_rolling['dt'], df_rolling['rolling_vol '+ 'ctrb_day'])
       plt.title("Rolling volatility " + str(window) + " days")
       plt.xticks(rotation=15)
       plt.savefig('rolling volatility_fund.png')
       plt.show()

       dframe_fund['Sum30'] = dframe_fund['ctrb_day'].rolling(30).sum()
       plt.plot(df_rolling['dt'], dframe_fund['Sum30'])
       skew1 = str(round(skew(dframe_fund['Sum30'].dropna()), 5))
       plt.title("Skewness for data " + skew1)
       plt.xticks(rotation=15)
       plt.savefig('Skewness for data.png')
       plt.show()

       ctrb_fund_1 = dframe_fund['ctrb_day'].tolist()
       ctrb_fund_1_array = np.asarray(ctrb_fund_1)

       plt.title("Autocorrelation Plot")
       plt.xlabel("Lags")
       plt.acorr(ctrb_fund_1_array, maxlags=None)
       plt.grid(True)
       plt.savefig('Autocorrelation_plot_2.png')
       plt.show()

       plot_acf(ctrb_fund_1_array).savefig('Autocorrelation_plot_1.png')

       wb = openpyxl.load_workbook(FilePath)
       wb.create_sheet('fund_portfolio_graph')
       active = wb['fund_portfolio_graph']
       active.add_image(Image('Frequency of return.png'), 'A1')
       active.add_image(Image('rolling volatility_fund.png'), 'K1')
       active.add_image(Image('Skewness for data.png'), 'K27')
       active.add_image(Image('Autocorrelation_plot_1.png'), 'U1')
       active.add_image(Image('Autocorrelation_plot_2.png'), 'U27')
       wb.save(FilePath)

       os.remove('Frequency of return.png')
       os.remove('rolling volatility_fund.png')
       os.remove('Skewness for data.png')
       os.remove('Autocorrelation_plot_1.png')
       os.remove('Autocorrelation_plot_2.png')

   def volatility_country():

       dframe_fund_1 = df_file.groupby(['dt','country']).agg({'ctrb_day': 'sum'}).reset_index()
       df_pivot = dframe_fund_1.pivot(index='dt', columns='country', values='ctrb_day').reset_index()
       col_countries = df_pivot.columns.tolist()
       col_countries.remove('dt')
       df_pivot=df_pivot.fillna(0)
       for index in range(len(col_countries)):
           volatility = count_volatility(df_pivot, str(col_countries[index]))
           df_info_vol[col_countries[index]]=volatility
           df_rolling = rolling_std_df(df_pivot, str(col_countries[index]), window)
           df_rolling.drop(col_countries[index], axis=1, inplace=True)

       FilePath = "volatility_fund " + File +".xlsx"
       ExcelWorkbook = load_workbook(FilePath)
       writer = pd.ExcelWriter(FilePath, engine='openpyxl')
       writer.book = ExcelWorkbook
       df_rolling.to_excel(writer, sheet_name='country', startrow=0, startcol=0)  ########
       df_info_vol.to_excel(writer, sheet_name='country', startrow=0, startcol=30)  ########
       writer.save()

       vol_year = df_info_vol.loc[4].tolist()
       vol_year.pop(0)

       plt.barh(col_countries, width=vol_year) # to clean
       plt.title("Volatility: country year (252 trading days)")
       plt.xlabel("Volatility")
       plt.savefig('volatility_country.png')
       plt.show()

       plt.barh(col_countries, width=vol_year)
       plt.yticks(rotation=15)
       plt.title("Volatility: country " + 'year (252 trading days)')
       plt.xlabel("Volatility")
       plt.savefig('volatility_country' + 'year (252 trading days)' + '.png')
       plt.show()

       wb = openpyxl.load_workbook(FilePath)
       wb.create_sheet('country_graph')
       active = wb['country_graph']
       active.add_image(Image('volatility_country' + volatility_name[4] + '.png'), 'A1')
       wb.save(FilePath)

       os.remove('volatility_country' + volatility_name[4] + '.png')

   def volatility_sector():
       unique_sector = pd.unique(df_file['sector'])

       volatility_name = ["daily", "month(21 trading days)", "3 month(63 trading days)", "6 month (63 trading days)", "year(252 trading days)"]
       for index in range(len(intervals) - 5):
           index = index + 5
           volatility_name.append(str(intervals[index]) + " trading days")
       df_info_vol = pd.DataFrame(volatility_name)

       dframe_fund_1 = df_file.groupby(['dt','sector']).agg({'ctrb_day': 'sum'}).reset_index()
       df_pivot = dframe_fund_1.pivot(index='dt', columns='sector', values='ctrb_day').reset_index()
       col_sector = df_pivot.columns.tolist()
       col_sector.remove('dt')
       df_pivot=df_pivot.fillna(0)
       for index in range(len(col_sector)):
           volatility = count_volatility(df_pivot, str(col_sector[index]))
           df_info_vol[col_sector[index]]=volatility
           df_rolling = rolling_std_df(df_pivot, str(col_sector[index]), window)
           df_rolling.drop(col_sector[index], axis=1, inplace=True)

       FilePath = "volatility_fund " + File +".xlsx"
       ExcelWorkbook = load_workbook(FilePath)
       writer = pd.ExcelWriter(FilePath, engine='openpyxl')
       writer.book = ExcelWorkbook
       df_rolling.to_excel(writer, sheet_name='sector', startrow=2, startcol=0)  ########
       df_info_vol.to_excel(writer, sheet_name='sector', startrow=2, startcol=20)  ########
       writer.save()

       vol_year = df_info_vol.loc[4].tolist()
       vol_year.pop(0)
       plt.barh(unique_sector, width=vol_year)
       plt.title("Volatility: sector " + volatility_name[4])
       plt.xlabel("Volatility")
       plt.savefig('volatility_sector ' + volatility_name[4] + '.png')
       plt.show()

       wb = openpyxl.load_workbook(FilePath)
       wb.create_sheet('sector_graph')
       active1 = wb['sector_graph']
       active1.add_image(Image('volatility_sector ' + volatility_name[4] + '.png'), 'A1')
       wb.save(FilePath)

       os.remove('volatility_sector ' + volatility_name[4] + '.png')

   def volatility_day_week():
       unique_day_week = pd.unique(df_file['day_week'])
       unique_day_week.sort()

       volatility_name = ["daily", "month (21 trading days)", "3 month (63 trading days)", "6 month (126 trading days)", "year(252 trading days)"]
       for index in range(len(intervals) - 5):
           index = index + 5
           volatility_name.append(str(intervals[index]) + " trading days")
       df_info_vol = pd.DataFrame(volatility_name)  ########

       dframe_fund_1 = df_file.groupby(['dt','day_week']).agg({'ctrb_day': 'sum'}).reset_index()
       df_pivot = dframe_fund_1.pivot(index='dt', columns='day_week', values='ctrb_day').reset_index()
       df_pivot = df_pivot.rename(columns={0: 'Monday', 1: 'Tuesday', 2: 'Wednesday', 3: 'Thursday', 4: 'Friday'})
       col_day_week = df_pivot.columns.tolist()
       col_day_week.remove('dt')
       df_pivot=df_pivot.fillna(0)
       for index in range(len(col_day_week)):
           volatility = count_volatility(df_pivot, str(col_day_week[index]))
           df_info_vol[col_day_week[index]]=volatility
           df_rolling = rolling_std_df(df_pivot, str(col_day_week[index]), window)
           df_rolling.drop(col_day_week[index], axis=1, inplace=True)

       FilePath = "volatility_fund " + File +".xlsx"
       ExcelWorkbook = load_workbook(FilePath)
       writer = pd.ExcelWriter(FilePath, engine='openpyxl')
       writer.book = ExcelWorkbook
       df_rolling.to_excel(writer, sheet_name='day_week', startrow=2, startcol=0)  ########
       df_info_vol.to_excel(writer, sheet_name='day_week', startrow=2, startcol=10)  ########
       writer.save()

       vol_year = df_info_vol.loc[4].tolist()
       vol_year.pop(0)
       plt.barh(unique_day_week, width=vol_year)
       plt.title("Volatility: day_week  " + volatility_name[4])
       plt.xlabel("Volatility")
       plt.savefig('volatility_day_week' + volatility_name[4] + '.png')
       plt.show()

       wb = openpyxl.load_workbook(FilePath)
       wb.create_sheet('day_week_graph')
       active1 = wb['day_week_graph']
       active1.add_image(Image('volatility_day_week' + volatility_name[4] + '.png'), 'A1')
       wb.save(FilePath)

       os.remove('volatility_day_week' + volatility_name[4] + '.png')

   def volatility_month():
       df_file['month'] = pd.to_datetime(df_file['dt']).dt.month_name()
       unique_month = pd.unique(df_file['month'])

       volatility_name = ["daily", "month(21 trading days)", "3 month(63 trading days)", "6 month (63 trading days)", "year(252 trading days)"]
       for index in range(len(intervals) - 5):
           index = index + 5
           volatility_name.append(str(intervals[index]) + " trading days")
       df_info_vol = pd.DataFrame(volatility_name)  ########

       dframe_fund_1 = df_file.groupby(['dt', 'month']).agg({'ctrb_day': 'sum'}).reset_index()
       df_pivot = dframe_fund_1.pivot(index='dt', columns='month', values='ctrb_day').reset_index()
       col_month = df_pivot.columns.tolist()
       col_month.remove('dt')
       df_pivot = df_pivot.fillna(0)
       for index in range(len(col_month)):
           volatility = count_volatility(df_pivot, str(col_month[index]))
           df_info_vol[col_month[index]]=volatility
           df_rolling = rolling_std_df(df_pivot, str(col_month[index]), window)
           df_rolling.drop(col_month[index], axis=1, inplace=True)

       FilePath = "volatility_fund " + File +".xlsx"
       ExcelWorkbook = load_workbook(FilePath)
       writer = pd.ExcelWriter(FilePath, engine='openpyxl')
       writer.book = ExcelWorkbook
       df_rolling.to_excel(writer, sheet_name='month', startrow=2, startcol=0)  ########
       df_info_vol.to_excel(writer, sheet_name='month', startrow=2, startcol=15)  ########
       writer.save()

       vol_year = df_info_vol.loc[4].tolist()
       vol_year.pop(0)
       plt.barh(unique_month, width=vol_year)
       plt.title("Volatility: month  " + volatility_name[4])
       plt.xlabel("Volatility")
       plt.savefig('volatility_month' + volatility_name[4] + '.png')
       plt.show()

       wb = openpyxl.load_workbook(FilePath)
       wb.create_sheet('month_graph')
       active1 = wb['month_graph']
       active1.add_image(Image('volatility_month' + volatility_name[4] + '.png'), 'A1')
       wb.save(FilePath)

       os.remove('volatility_month' + volatility_name[4] + '.png')

   volatility_portfolio()
   volatility_country()
   volatility_sector()
   volatility_day_week()
   volatility_month()


intervals = [1, 21, 63, 126, 252]
window = 60
list = ['2022.xlsx', '2021.xlsx', '2020.xlsx', '2019.xlsx', '2018.xlsx', '2017.xlsx']

fund_years(list, window)
for index in range(len(list)):
    fund_year(list[index])

